import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

file_name = "result.xlsx"

# 🎨 Colors
green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# ==============================
# ✅ LOAD OR CREATE
# ==============================
if os.path.exists(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Test Case",
        "Form Name",
        "Website",
        "Execution Time",
        "Status",
        "Screenshot"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)


# ==============================
# ✅ ADD NEW COLUMNS DYNAMICALLY
# ==============================
def update_headers(data):
    global headers

    new_keys = list(data.keys())

    for key in new_keys:
        if key not in headers:
            headers.append(key)
            ws.cell(row=1, column=len(headers)).value = key


# ==============================
# ✅ GET TC ID
# ==============================
def get_next_tc(form_name):
    return f"{form_name.upper()}_TC_{ws.max_row}"


# ==============================
# ✅ WRITE RESULT
# ==============================
def write_result(form_name, website, status, screenshot, data):
    update_headers(data)

    row_data = [""] * len(headers)

    # Fixed columns
    row_data[headers.index("Test Case")] = get_next_tc(form_name)
    row_data[headers.index("Form Name")] = form_name
    row_data[headers.index("Website")] = website
    row_data[headers.index("Execution Time")] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_data[headers.index("Status")] = status
    row_data[headers.index("Screenshot")] = screenshot

    # Dynamic JSON fields
    for key, value in data.items():
        if key in headers:
            row_data[headers.index(key)] = value

    ws.append(row_data)

    # 🎨 Color status
    row = ws.max_row
    status_cell = ws.cell(row=row, column=headers.index("Status") + 1)

    if status == "PASS":
        status_cell.fill = green
    else:
        status_cell.fill = red


# ==============================
# ✅ AUTO WIDTH
# ==============================
def auto_adjust():
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3


# ==============================
# ✅ SAVE
# ==============================
def save():
    auto_adjust()
    wb.save(file_name)